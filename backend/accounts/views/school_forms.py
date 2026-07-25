import io
import calendar
import logging
from datetime import date, datetime

from django.db import transaction
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ..models.school_forms import (
    SchoolForm1, SchoolForm1Student,
    SchoolForm5, SchoolForm5Student,
    SchoolForm9, SchoolForm9Subject,
    SchoolForm10, SchoolForm10Record, SchoolForm10Subject,
)
from ..models.academic import Classroom, StudentClassEnrollment, SystemSetting, Subject
from ..models.user import User, Profile
from ..serializers.school_forms import (
    SchoolForm1ListSerializer, SchoolForm1DetailSerializer,
    SchoolForm1StudentSerializer, GenerateSF1Serializer,
    SF2OverviewSerializer,
    SchoolForm5ListSerializer, SchoolForm5DetailSerializer, GenerateSF5Serializer,
    SchoolForm9ListSerializer, SchoolForm9DetailSerializer, GenerateSF9Serializer,
    SchoolForm10ListSerializer, SchoolForm10DetailSerializer, GenerateSF10Serializer,
)
from ..utils import log_audit_action

logger = logging.getLogger(__name__)


class SchoolForm1ViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        'students__student__last_name', 'students__student__first_name',
        'students__student__profile__lrn', 'section', 'school_year',
        'grade_level', 'adviser__last_name', 'adviser__first_name',
    ]
    ordering_fields = ['school_year', 'grade_level', 'section', 'generated_at', 'status']
    ordering = ['-generated_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return SchoolForm1ListSerializer
        if self.action in ('retrieve', 'generate'):
            return SchoolForm1DetailSerializer
        return SchoolForm1ListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = SchoolForm1.objects.select_related(
            'adviser', 'generated_by',
        ).prefetch_related('students__student__profile', 'students__enrollment')

        # Role-based filtering — students and parents see limited records
        if user.role == 'student':
            qs = qs.filter(students__student=user)
        elif user.role == 'parent':
            linked = user.profile.linked_students.all() if hasattr(user, 'profile') else User.objects.none()
            qs = qs.filter(students__student__in=linked)

        # Filters
        school_year = self.request.query_params.get('school_year')
        grade_level = self.request.query_params.get('grade_level')
        section = self.request.query_params.get('section')
        adviser = self.request.query_params.get('adviser')
        sf_status = self.request.query_params.get('status')

        if school_year:
            qs = qs.filter(school_year=school_year)
        if grade_level:
            qs = qs.filter(grade_level=grade_level)
        if section:
            qs = qs.filter(section__icontains=section)
        if adviser:
            qs = qs.filter(
                Q(adviser__first_name__icontains=adviser) |
                Q(adviser__last_name__icontains=adviser)
            )
        if sf_status:
            qs = qs.filter(status=sf_status)

        return qs.distinct()

    def perform_destroy(self, instance):
        if instance.status == 'final':
            raise ValueError('Cannot delete a finalized SF1. Archive it instead.')
        log_audit_action(
            user=self.request.user,
            action='delete',
            model_name='SchoolForm1',
            object_id=instance.id,
            object_repr=str(instance),
            description=f'Deleted SF1: {instance}',
            request=self.request,
        )
        instance.delete()

    @action(detail=False, methods=['post'])
    def generate(self, request):
        serializer = GenerateSF1Serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        sf1 = serializer.save()

        log_audit_action(
            user=request.user,
            action='generate',
            model_name='SchoolForm1',
            object_id=sf1.id,
            object_repr=str(sf1),
            description=f'Generated SF1: {sf1} with {sf1.total_learners} students',
            request=request,
        )

        return Response(
            SchoolForm1DetailSerializer(sf1).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['post'])
    def regenerate(self, request, pk=None):
        sf1 = self.get_object()
        data = {
            'academic_year': sf1.school_year,
            'grade_level': sf1.grade_level,
            'section': sf1.section,
            'regenerate': True,
        }
        serializer = GenerateSF1Serializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        new_sf1 = serializer.save()

        log_audit_action(
            user=request.user,
            action='regenerate',
            model_name='SchoolForm1',
            object_id=new_sf1.id,
            object_repr=str(new_sf1),
            description=f'Regenerated SF1: {new_sf1} with {new_sf1.total_learners} students',
            request=request,
        )

        return Response(SchoolForm1DetailSerializer(new_sf1).data)

    @action(detail=True, methods=['put'])
    def update_status(self, request, pk=None):
        sf1 = self.get_object()
        new_status = request.data.get('status')
        if new_status not in ('draft', 'final', 'archived'):
            return Response({'error': 'Invalid status'}, status=400)

        sf1.status = new_status
        sf1.save(update_fields=['status'])

        log_audit_action(
            user=request.user,
            action='update_status',
            model_name='SchoolForm1',
            object_id=sf1.id,
            object_repr=str(sf1),
            description=f'Changed SF1 status to {new_status}',
            request=request,
        )

        return Response(SchoolForm1DetailSerializer(sf1).data)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        sf1 = self.get_object()
        students = sf1.students.select_related(
            'student', 'student__profile', 'enrollment',
        ).order_by('order')

        try:
            pdf_bytes = _generate_sf1_pdf(sf1, students)
        except ImportError:
            return Response(
                {'error': 'PDF generation requires reportlab. Install with: pip install reportlab'},
                status=500,
            )

        log_audit_action(
            user=request.user,
            action='export_pdf',
            model_name='SchoolForm1',
            object_id=sf1.id,
            object_repr=str(sf1),
            description=f'Exported PDF for SF1: {sf1}',
            request=request,
        )

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f"SF1_{sf1.school_year}_{sf1.grade_level}_{sf1.section}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        sf1 = self.get_object()
        students = sf1.students.select_related(
            'student', 'student__profile', 'enrollment',
        ).order_by('order')

        try:
            xlsx_bytes = _generate_sf1_excel(sf1, students)
        except ImportError:
            return Response(
                {'error': 'Excel generation requires openpyxl. Install with: pip install openpyxl'},
                status=500,
            )

        log_audit_action(
            user=request.user,
            action='export_excel',
            model_name='SchoolForm1',
            object_id=sf1.id,
            object_repr=str(sf1),
            description=f'Exported Excel for SF1: {sf1}',
            request=request,
        )

        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"SF1_{sf1.school_year}_{sf1.grade_level}_{sf1.section}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'])
    def print_view(self, request, pk=None):
        sf1 = self.get_object()
        students = sf1.students.select_related(
            'student', 'student__profile', 'enrollment',
        ).order_by('order')

        try:
            pdf_bytes = _generate_sf1_pdf(sf1, students)
        except ImportError:
            return Response(
                {'error': 'PDF generation requires reportlab.'},
                status=500,
            )

        log_audit_action(
            user=request.user,
            action='print',
            model_name='SchoolForm1',
            object_id=sf1.id,
            object_repr=str(sf1),
            description=f'Printed SF1: {sf1}',
            request=request,
        )

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="SF1_{sf1.school_year}_{sf1.grade_level}_{sf1.section}.pdf"'
        return response


def _get_student_field(student, field, default=''):
    try:
        profile = student.profile
        return getattr(profile, field, default) or default
    except (Profile.DoesNotExist, AttributeError):
        return default


def _calculate_age(dob):
    if not dob:
        return ''
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def _generate_sf1_pdf(sf1, students):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LEGAL, landscape
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(LEGAL),
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        topMargin=0.3 * inch, bottomMargin=0.3 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=11, spaceAfter=2)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=8, alignment=1)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=7, leading=9, fontName='Helvetica-Bold')

    elements = []

    # Header
    settings = SystemSetting.get_settings()
    elements.append(Paragraph("Republic of the Philippines", subtitle_style))
    elements.append(Paragraph("Department of Education", subtitle_style))
    elements.append(Paragraph(f"Region: {getattr(settings, 'region', '')} | Division: {getattr(settings, 'division', '')}", subtitle_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"<b>{settings.site_name}</b>", title_style))
    elements.append(Paragraph(f"School ID: {getattr(settings, 'school_id', '')}", subtitle_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"<b>SCHOOL FORM 1 (SF1) — School Register</b>  |  "
        f"School Year: {sf1.school_year}  |  Grade: {sf1.grade_level}  |  "
        f"Section: {sf1.section}  |  Class Adviser: {sf1.adviser.profile.title if sf1.adviser and hasattr(sf1.adviser, 'profile') else ''} "
        f"{sf1.adviser.first_name} {sf1.adviser.last_name if sf1.adviser else ''}",
        ParagraphStyle('FormTitle', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=1, spaceAfter=6),
    ))

    # Table header
    headers = [
        Paragraph('<b>No.</b>', cell_style),
        Paragraph('<b>LRN</b>', cell_style),
        Paragraph('<b>Last Name</b>', cell_style),
        Paragraph('<b>First Name</b>', cell_style),
        Paragraph('<b>Middle Name</b>', cell_style),
        Paragraph('<b>Ext.</b>', cell_style),
        Paragraph('<b>Sex</b>', cell_style),
        Paragraph('<b>Date of Birth</b>', cell_style),
        Paragraph('<b>Age</b>', cell_style),
        Paragraph('<b>Mother Tongue</b>', cell_style),
        Paragraph('<b>IP</b>', cell_style),
        Paragraph('<b>Religion</b>', cell_style),
        Paragraph('<b>Address</b>', cell_style),
        Paragraph('<b>Parent/Guardian</b>', cell_style),
        Paragraph('<b>Contact</b>', cell_style),
        Paragraph('<b>Status</b>', cell_style),
    ]

    data = [headers]

    for idx, entry in enumerate(students, 1):
        s = entry.student
        profile = s.profile if hasattr(s, 'profile') else None
        dob = profile.date_of_birth if profile else None
        parent_name = (profile.mother_name or profile.father_name or '') if profile else ''

        row = [
            Paragraph(str(idx), cell_style),
            Paragraph(_get_student_field(s, 'lrn'), cell_style),
            Paragraph(s.last_name or '', cell_style),
            Paragraph(s.first_name or '', cell_style),
            Paragraph(_get_student_field(s, 'middle_name'), cell_style),
            Paragraph(_get_student_field(s, 'extension_name'), cell_style),
            Paragraph((_get_student_field(s, 'sex')).title(), cell_style),
            Paragraph(dob.strftime('%m/%d/%Y') if dob else '', cell_style),
            Paragraph(str(_calculate_age(dob)) if dob else '', cell_style),
            Paragraph(_get_student_field(s, 'mother_tongue'), cell_style),
            Paragraph(_get_student_field(s, 'indigenous_people'), cell_style),
            Paragraph(_get_student_field(s, 'religion'), cell_style),
            Paragraph(_get_student_field(s, 'address'), cell_style),
            Paragraph(parent_name, cell_style),
            Paragraph(_get_student_field(s, 'phone_number'), cell_style),
            Paragraph(entry.enrollment_status if hasattr(entry, 'enrollment_status') else 'Enrolled', cell_style),
        ]
        data.append(row)

    # Summary row
    summary = [
        '', '', '', '', '', '', '', '', '', '', '', '',
        Paragraph(f'<b>Total Male: {sf1.total_male}</b>', cell_style),
        Paragraph(f'<b>Total Female: {sf1.total_female}</b>', cell_style),
        Paragraph(f'<b>Total: {sf1.total_learners}</b>', cell_style),
        '',
    ]
    data.append(summary)

    col_widths = [0.3*inch, 0.7*inch, 0.75*inch, 0.75*inch, 0.65*inch, 0.35*inch, 0.35*inch, 0.6*inch, 0.3*inch, 0.6*inch, 0.45*inch, 0.55*inch, 1.1*inch, 0.9*inch, 0.65*inch, 0.55*inch]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D1B4D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
    ]))

    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def _generate_sf1_excel(sf1, students):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"SF1 {sf1.grade_level} {sf1.section}"

    settings = SystemSetting.get_settings()

    # Title rows
    ws.merge_cells('A1:P1')
    ws['A1'] = 'Republic of the Philippines'
    ws['A1'].font = Font(size=9, italic=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:P2')
    ws['A2'] = 'Department of Education'
    ws['A2'].font = Font(size=9, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:P3')
    ws['A3'] = f"Region: {getattr(settings, 'region', '')} | Division: {getattr(settings, 'division', '')}"
    ws['A3'].font = Font(size=8)
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:P4')
    ws['A4'] = settings.site_name
    ws['A4'].font = Font(size=12, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A5:P5')
    ws['A5'] = f"School ID: {getattr(settings, 'school_id', '')}"
    ws['A5'].font = Font(size=8)
    ws['A5'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A6:P6')
    ws['A6'] = (
        f"SCHOOL FORM 1 (SF1) — School Register  |  "
        f"School Year: {sf1.school_year}  |  Grade: {sf1.grade_level}  |  "
        f"Section: {sf1.section}  |  Class Adviser: "
        f"{sf1.adviser.profile.title if sf1.adviser and hasattr(sf1.adviser, 'profile') else ''} "
        f"{sf1.adviser.first_name} {sf1.adviser.last_name if sf1.adviser else ''}"
    )
    ws['A6'].font = Font(size=9, bold=True)
    ws['A6'].alignment = Alignment(horizontal='center')

    # Header row
    headers = [
        'No.', 'LRN', 'Last Name', 'First Name', 'Middle Name', 'Ext.',
        'Sex', 'Date of Birth', 'Age', 'Mother Tongue', 'IP',
        'Religion', 'Address', 'Parent/Guardian', 'Contact', 'Status',
    ]

    header_fill = PatternFill(start_color='2D1B4D', end_color='2D1B4D', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=8)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    data_font = Font(size=8)
    alt_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')

    for idx, entry in enumerate(students, 1):
        s = entry.student
        profile = s.profile if hasattr(s, 'profile') else None
        dob = profile.date_of_birth if profile else None
        parent_name = (profile.mother_name or profile.father_name or '') if profile else ''

        row_data = [
            idx,
            _get_student_field(s, 'lrn'),
            s.last_name or '',
            s.first_name or '',
            _get_student_field(s, 'middle_name'),
            _get_student_field(s, 'extension_name'),
            (_get_student_field(s, 'sex')).title(),
            dob.strftime('%m/%d/%Y') if dob else '',
            _calculate_age(dob) if dob else '',
            _get_student_field(s, 'mother_tongue'),
            _get_student_field(s, 'indigenous_people'),
            _get_student_field(s, 'religion'),
            _get_student_field(s, 'address'),
            parent_name,
            _get_student_field(s, 'phone_number'),
            'Enrolled',
        ]

        row_num = idx + 7
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = alt_fill

    # Summary row
    summary_row = len(students) + 8
    ws.cell(row=summary_row, column=13, value=f"Total Male: {sf1.total_male}").font = Font(bold=True, size=8)
    ws.cell(row=summary_row, column=14, value=f"Total Female: {sf1.total_female}").font = Font(bold=True, size=8)
    ws.cell(row=summary_row, column=15, value=f"Total: {sf1.total_learners}").font = Font(bold=True, size=8)

    # Column widths
    widths = [5, 15, 15, 15, 12, 6, 6, 12, 5, 12, 10, 10, 25, 18, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# SF2 — Daily Attendance Report of Learners
# ═══════════════════════════════════════════════════════════════════════════════

def _find_classroom(data):
    classroom = Classroom.objects.filter(
        grade_level=data['grade_level'], name__icontains=data['section'],
    ).first()
    if not classroom:
        try:
            from portal.models import AcademicYear
            ay = AcademicYear.objects.filter(name=data['academic_year']).first()
            if ay:
                classroom = Classroom.objects.filter(
                    grade_level=data['grade_level'], academic_year=ay,
                ).filter(Q(name__icontains=data['section']) | Q(name__iexact=data['section'])).first()
        except Exception:
            pass
    return classroom


class SF2ViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def overview(self, request):
        academic_year = request.query_params.get('academic_year', SystemSetting.get_settings().academic_year)
        grade_level = request.query_params.get('grade_level', '')
        section = request.query_params.get('section', '')
        month = request.query_params.get('month', str(timezone.now().month))
        year = request.query_params.get('year', str(timezone.now().year))

        if not grade_level or not section:
            return Response({'error': 'grade_level and section are required'}, status=400)

        classroom = _find_classroom({'academic_year': academic_year, 'grade_level': grade_level, 'section': section})
        if not classroom:
            return Response({'error': 'Classroom not found'}, status=404)

        month_int = int(month)
        year_int = int(year)
        _, num_days = calendar.monthrange(year_int, month_int)

        students = User.objects.filter(
            enrollments__classroom=classroom, role='student',
        ).select_related('profile').order_by('last_name', 'first_name')

        from ..models.attendance import Attendance

        student_rows = []
        total_present = total_absent = total_late = 0
        school_days = 0

        for day in range(1, num_days + 1):
            d = date(year_int, month_int, day)
            if d.weekday() >= 5:
                continue
            if d > date.today():
                continue
            school_days += 1

        for student in students:
            daily = {}
            p_count = a_count = l_count = 0
            for day in range(1, num_days + 1):
                d = date(year_int, month_int, day)
                if d.weekday() >= 5 or d > date.today():
                    continue
                att = Attendance.objects.filter(student=student, classroom=classroom, date=d).first()
                status_code = '—'
                if att:
                    if att.status == 'present':
                        status_code = 'P'
                        p_count += 1
                    elif att.status == 'absent':
                        status_code = 'A'
                        a_count += 1
                    elif att.status == 'late':
                        status_code = 'L'
                        l_count += 1
                    elif att.status == 'excused':
                        status_code = 'E'
                        p_count += 1
                daily[str(day)] = status_code

            total_present += p_count
            total_absent += a_count
            total_late += l_count
            attended = p_count + l_count
            pct = round((attended / school_days * 100), 1) if school_days > 0 else 0

            p = _get_profile(student) if False else None
            try:
                p = student.profile
            except Exception:
                p = None

            student_rows.append({
                'student_id': student.id,
                'student_name': full_name(student),
                'lrn': p.lrn if p else '',
                'sex': (p.sex or '').title() if p else '',
                'daily': daily,
                'days_present': p_count,
                'days_absent': a_count,
                'days_late': l_count,
                'attendance_pct': pct,
            })

        total_enrolled = len(student_rows)
        overall_pct = round((total_present / (school_days * total_enrolled) * 100), 1) if school_days > 0 and total_enrolled > 0 else 0

        settings = SystemSetting.get_settings()
        adviser_name = full_name(classroom.teacher) if classroom.teacher else ''

        return Response({
            'school_name': settings.site_name,
            'school_id': getattr(settings, 'school_id', ''),
            'region': getattr(settings, 'region', ''),
            'division': getattr(settings, 'division', ''),
            'school_year': academic_year,
            'grade_level': grade_level,
            'section': section,
            'adviser_name': adviser_name,
            'total_learners': total_enrolled,
            'overall_attendance_pct': overall_pct,
            'month': month_int,
            'year': year_int,
            'month_name': calendar.month_name[month_int],
            'total_school_days': school_days,
            'students': student_rows,
            'total_present': total_present,
            'total_absent': total_absent,
            'total_late': total_late,
        })

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        data = self._get_overview_data(request)
        if isinstance(data, Response):
            return data
        try:
            pdf_bytes = _generate_sf2_pdf(data)
        except ImportError:
            return Response({'error': 'reportlab required'}, status=500)
        log_audit_action(user=request.user, action='export_pdf', model_name='SF2',
                         description=f'Exported SF2 PDF for {data["grade_level"]} {data["section"]}',
                         request=request)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="SF2_{data["school_year"]}_{data["grade_level"]}_{data["section"]}_{data["month_name"]}.pdf"'
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        data = self._get_overview_data(request)
        if isinstance(data, Response):
            return data
        try:
            xlsx_bytes = _generate_sf2_excel(data)
        except ImportError:
            return Response({'error': 'openpyxl required'}, status=500)
        log_audit_action(user=request.user, action='export_excel', model_name='SF2',
                         description=f'Exported SF2 Excel for {data["grade_level"]} {data["section"]}',
                         request=request)
        response = HttpResponse(xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="SF2_{data["school_year"]}_{data["grade_level"]}_{data["section"]}_{data["month_name"]}.xlsx"'
        return response

    def _get_overview_data(self, request):
        resp = self.overview(request)
        if resp.status_code != 200:
            return resp
        return resp.data


def _generate_sf2_pdf(data):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LEGAL, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(LEGAL),
        leftMargin=0.3*inch, rightMargin=0.3*inch, topMargin=0.3*inch, bottomMargin=0.3*inch)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('C', parent=styles['Normal'], fontSize=6, leading=7)
    hs = ParagraphStyle('H', parent=styles['Normal'], fontSize=6, leading=7, fontName='Helvetica-Bold')
    ts = ParagraphStyle('T', parent=styles['Title'], fontSize=10)
    ss = ParagraphStyle('S', parent=styles['Normal'], fontSize=7, alignment=1)

    elements = []
    settings = SystemSetting.get_settings()
    elements.append(Paragraph("Republic of the Philippines", ss))
    elements.append(Paragraph("Department of Education", ss))
    elements.append(Paragraph(f"<b>{data['school_name']}</b>", ts))
    elements.append(Paragraph(f"School ID: {data['school_id']}", ss))
    elements.append(Spacer(1, 3))
    elements.append(Paragraph(
        f"<b>SCHOOL FORM 2 (SF2) — Daily Attendance Report of Learners</b>  |  "
        f"SY: {data['school_year']}  |  Grade: {data['grade_level']}  |  "
        f"Section: {data['section']}  |  Month: {data['month_name']} {data['year']}  |  "
        f"Adviser: {data['adviser_name']}",
        ParagraphStyle('FT', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=1, spaceAfter=4)))

    num_days = calendar.monthrange(data['year'], data['month'])[1]
    headers = [Paragraph('<b>No.</b>', hs), Paragraph('<b>Name of Learner</b>', hs)]
    for d in range(1, num_days + 1):
        dt = date(data['year'], data['month'], d)
        if dt.weekday() < 5 and dt <= date.today():
            headers.append(Paragraph(f'<b>{d}</b>', hs))
    headers += [Paragraph('<b>Days Present</b>', hs), Paragraph('<b>Days Absent</b>', hs),
                Paragraph('<b>Days Late</b>', hs), Paragraph('<b>%</b>', hs)]

    table_data = [headers]
    for idx, s in enumerate(data['students'], 1):
        row = [Paragraph(str(idx), cs), Paragraph(s['student_name'], cs)]
        for d in range(1, num_days + 1):
            dt = date(data['year'], data['month'], d)
            if dt.weekday() < 5 and dt <= date.today():
                row.append(Paragraph(s['daily'].get(str(d), '—'), cs))
        row += [Paragraph(str(s['days_present']), cs), Paragraph(str(s['days_absent']), cs),
                Paragraph(str(s['days_late']), cs), Paragraph(f"{s['attendance_pct']}%", cs)]
        table_data.append(row)

    summary_row = [Paragraph(f'<b>Total: {data["total_learners"]}</b>', hs), '']
    for d in range(1, num_days + 1):
        dt = date(data['year'], data['month'], d)
        if dt.weekday() < 5 and dt <= date.today():
            summary_row.append('')
    summary_row += [Paragraph(f'<b>{data["total_present"]}</b>', hs),
                    Paragraph(f'<b>{data["total_absent"]}</b>', hs),
                    Paragraph(f'<b>{data["total_late"]}</b>', hs),
                    Paragraph(f'<b>{data["overall_attendance_pct"]}%</b>', hs)]
    table_data.append(summary_row)

    col_widths = [0.3*inch, 1.5*inch] + [0.28*inch] * (len(headers) - 4) + [0.5*inch, 0.5*inch, 0.4*inch, 0.4*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D1B4D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def _generate_sf2_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"SF2 {data['month_name']}"
    num_days = calendar.monthrange(data['year'], data['month'])[1]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 5)
    ws.cell(1, 1, f"SF2 — Daily Attendance Report | {data['school_name']} | {data['grade_level']} {data['section']} | {data['month_name']} {data['year']}").font = Font(bold=True, size=10)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    headers = ['No.', 'Name of Learner']
    for d in range(1, num_days + 1):
        dt = date(data['year'], data['month'], d)
        if dt.weekday() < 5 and dt <= date.today():
            headers.append(str(d))
    headers += ['Present', 'Absent', 'Late', '%']

    hdr_fill = PatternFill(start_color='2D1B4D', end_color='2D1B4D', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=7)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')

    for idx, s in enumerate(data['students'], 1):
        r = idx + 3
        ws.cell(r, 1, idx).border = thin
        ws.cell(r, 2, s['student_name']).border = thin
        col = 3
        for d in range(1, num_days + 1):
            dt = date(data['year'], data['month'], d)
            if dt.weekday() < 5 and dt <= date.today():
                cell = ws.cell(r, col, s['daily'].get(str(d), ''))
                cell.border = thin
                cell.alignment = Alignment(horizontal='center')
                col += 1
        ws.cell(r, col, s['days_present']).border = thin
        ws.cell(r, col+1, s['days_absent']).border = thin
        ws.cell(r, col+2, s['days_late']).border = thin
        ws.cell(r, col+3, f"{s['attendance_pct']}%").border = thin

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# SF5 — Report on Promotion and Learning Progress
# ═══════════════════════════════════════════════════════════════════════════════

class SchoolForm5ViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['students__student__last_name', 'students__student__first_name',
                     'section', 'school_year', 'grade_level']
    ordering_fields = ['school_year', 'grade_level', 'generated_at', 'status']
    ordering = ['-generated_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return SchoolForm5ListSerializer
        return SchoolForm5DetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = SchoolForm5.objects.select_related('adviser', 'generated_by').prefetch_related('students__student__profile')
        if user.role == 'student':
            qs = qs.filter(students__student=user)
        elif user.role == 'parent':
            linked = user.profile.linked_students.all() if hasattr(user, 'profile') else User.objects.none()
            qs = qs.filter(students__student__in=linked)

        for param in ['school_year', 'grade_level', 'status']:
            val = self.request.query_params.get(param)
            if val:
                qs = qs.filter(**{param: val})
        section = self.request.query_params.get('section')
        if section:
            qs = qs.filter(section__icontains=section)
        return qs.distinct()

    def perform_destroy(self, instance):
        if instance.status == 'final':
            raise ValueError('Cannot delete a finalized SF5.')
        log_audit_action(user=self.request.user, action='delete', model_name='SchoolForm5',
                         object_id=instance.id, object_repr=str(instance),
                         description=f'Deleted SF5: {instance}', request=self.request)
        instance.delete()

    @action(detail=False, methods=['post'])
    def generate(self, request):
        serializer = GenerateSF5Serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        sf5 = serializer.save()
        log_audit_action(user=request.user, action='generate', model_name='SchoolForm5',
                         object_id=sf5.id, object_repr=str(sf5),
                         description=f'Generated SF5: {sf5} with {sf5.total_learners} students',
                         request=request)
        return Response(SchoolForm5DetailSerializer(sf5).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def regenerate(self, request, pk=None):
        sf5 = self.get_object()
        data = {'academic_year': sf5.school_year, 'grade_level': sf5.grade_level,
                'section': sf5.section, 'regenerate': True}
        serializer = GenerateSF5Serializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        new_sf5 = serializer.save()
        log_audit_action(user=request.user, action='regenerate', model_name='SchoolForm5',
                         object_id=new_sf5.id, object_repr=str(new_sf5),
                         description=f'Regenerated SF5: {new_sf5}', request=request)
        return Response(SchoolForm5DetailSerializer(new_sf5).data)

    @action(detail=True, methods=['put'])
    def update_status(self, request, pk=None):
        sf5 = self.get_object()
        new_status = request.data.get('status')
        if new_status not in ('draft', 'final', 'archived'):
            return Response({'error': 'Invalid status'}, status=400)
        sf5.status = new_status
        sf5.save(update_fields=['status'])
        log_audit_action(user=request.user, action='update_status', model_name='SchoolForm5',
                         object_id=sf5.id, object_repr=str(sf5),
                         description=f'Changed SF5 status to {new_status}', request=request)
        return Response(SchoolForm5DetailSerializer(sf5).data)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        sf5 = self.get_object()
        try:
            pdf_bytes = _generate_sf5_pdf(sf5)
        except ImportError:
            return Response({'error': 'reportlab required'}, status=500)
        log_audit_action(user=request.user, action='export_pdf', model_name='SchoolForm5',
                         object_id=sf5.id, description=f'Exported SF5 PDF: {sf5}', request=request)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="SF5_{sf5.school_year}_{sf5.grade_level}_{sf5.section}.pdf"'
        return response

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        sf5 = self.get_object()
        try:
            xlsx_bytes = _generate_sf5_excel(sf5)
        except ImportError:
            return Response({'error': 'openpyxl required'}, status=500)
        log_audit_action(user=request.user, action='export_excel', model_name='SchoolForm5',
                         object_id=sf5.id, description=f'Exported SF5 Excel: {sf5}', request=request)
        response = HttpResponse(xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="SF5_{sf5.school_year}_{sf5.grade_level}_{sf5.section}.xlsx"'
        return response


def _generate_sf5_pdf(sf5):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LEGAL
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LEGAL,
        leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('C', parent=styles['Normal'], fontSize=7, leading=9)
    ss = ParagraphStyle('S', parent=styles['Normal'], fontSize=8, alignment=1)

    elements = []
    settings = SystemSetting.get_settings()
    elements.append(Paragraph("Republic of the Philippines", ss))
    elements.append(Paragraph("Department of Education", ss))
    elements.append(Paragraph(f"<b>{settings.site_name}</b>", ParagraphStyle('T', parent=styles['Title'], fontSize=11)))
    elements.append(Paragraph(
        f"<b>SCHOOL FORM 5 (SF5) — Report on Promotion and Learning Progress</b>  |  "
        f"SY: {sf5.school_year}  |  Grade: {sf5.grade_level}  |  Section: {sf5.section}  |  "
        f"Adviser: {full_name(sf5.adviser) if sf5.adviser else ''}",
        ParagraphStyle('FT', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=1, spaceAfter=6)))

    students = sf5.students.select_related('student', 'student__profile').order_by('order')
    headers = ['No.', 'LRN', 'Name', 'Sex', 'Gen. Avg', 'Passed', 'Failed', 'Status', 'Awards', 'Remarks']
    table_data = [[Paragraph(f'<b>{h}</b>', cs) for h in headers]]

    for idx, s in enumerate(students, 1):
        p = s.student.profile if hasattr(s.student, 'profile') else None
        table_data.append([
            Paragraph(str(idx), cs),
            Paragraph(p.lrn if p else '', cs),
            Paragraph(full_name(s.student), cs),
            Paragraph((p.sex or '').title() if p else '', cs),
            Paragraph(str(s.general_average) if s.general_average else '—', cs),
            Paragraph(str(s.passed_subjects), cs),
            Paragraph(str(s.failed_subjects), cs),
            Paragraph(s.get_promotion_status_display(), cs),
            Paragraph(s.awards, cs),
            Paragraph(s.remarks, cs),
        ])

    summary = [
        '', '', Paragraph(f'<b>Total: {sf5.total_learners}</b>', cs), '', '',
        Paragraph(f'<b>Promoted: {sf5.total_promoted}</b>', cs),
        Paragraph(f'<b>Retained: {sf5.total_retained}</b>', cs),
        Paragraph(f'<b>Conditional: {sf5.total_conditional}</b>', cs), '', '',
    ]
    table_data.append(summary)

    col_widths = [0.35*inch, 0.8*inch, 1.5*inch, 0.4*inch, 0.55*inch, 0.45*inch, 0.45*inch, 1.0*inch, 1.0*inch, 1.2*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D1B4D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def _generate_sf5_excel(sf5):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "SF5"
    headers = ['No.', 'LRN', 'Last Name', 'First Name', 'Sex', 'Gen. Average', 'Passed', 'Failed', 'Status', 'Awards', 'Remarks']
    hdr_fill = PatternFill(start_color='2D1B4D', end_color='2D1B4D', fill_type='solid')
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color='FFFFFF', size=8)
        cell.fill = hdr_fill
        cell.border = thin

    students = sf5.students.select_related('student', 'student__profile').order_by('order')
    for idx, s in enumerate(students, 1):
        p = s.student.profile if hasattr(s.student, 'profile') else None
        r = idx + 1
        ws.cell(r, 1, idx).border = thin
        ws.cell(r, 2, p.lrn if p else '').border = thin
        ws.cell(r, 3, s.student.last_name or '').border = thin
        ws.cell(r, 4, s.student.first_name or '').border = thin
        ws.cell(r, 5, (p.sex or '').title() if p else '').border = thin
        ws.cell(r, 6, float(s.general_average) if s.general_average else '').border = thin
        ws.cell(r, 7, s.passed_subjects).border = thin
        ws.cell(r, 8, s.failed_subjects).border = thin
        ws.cell(r, 9, s.get_promotion_status_display()).border = thin
        ws.cell(r, 10, s.awards).border = thin
        ws.cell(r, 11, s.remarks).border = thin

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# SF9 — Learner's Progress Report Card
# ═══════════════════════════════════════════════════════════════════════════════

class SchoolForm9ViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['student__last_name', 'student__first_name', 'student__profile__lrn',
                     'section', 'school_year', 'grade_level']
    ordering_fields = ['school_year', 'grade_level', 'generated_at', 'status']
    ordering = ['-generated_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return SchoolForm9ListSerializer
        return SchoolForm9DetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = SchoolForm9.objects.select_related('student', 'adviser', 'student__profile').prefetch_related('subjects__subject')
        if user.role == 'student':
            qs = qs.filter(student=user)
        elif user.role == 'parent':
            linked = user.profile.linked_students.all() if hasattr(user, 'profile') else User.objects.none()
            qs = qs.filter(student__in=linked)
        elif user.role == 'staff':
            qs = qs.filter(adviser=user)

        for param in ['school_year', 'grade_level', 'status']:
            val = self.request.query_params.get(param)
            if val:
                qs = qs.filter(**{param: val})
        section = self.request.query_params.get('section')
        if section:
            qs = qs.filter(section__icontains=section)
        student_id = self.request.query_params.get('student')
        if student_id:
            qs = qs.filter(student_id=student_id)
        return qs.distinct()

    def perform_destroy(self, instance):
        log_audit_action(user=self.request.user, action='delete', model_name='SchoolForm9',
                         object_id=instance.id, object_repr=str(instance),
                         description=f'Deleted SF9: {instance}', request=self.request)
        instance.delete()

    @action(detail=False, methods=['post'])
    def generate(self, request):
        serializer = GenerateSF9Serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        records = serializer.save()
        log_audit_action(user=request.user, action='generate', model_name='SchoolForm9',
                         description=f'Generated {len(records)} SF9 report cards', request=request)
        return Response(SchoolForm9DetailSerializer(records, many=True).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def batch_generate(self, request):
        serializer = GenerateSF9Serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        records = serializer.save()
        log_audit_action(user=request.user, action='batch_generate', model_name='SchoolForm9',
                         description=f'Batch generated {len(records)} SF9 cards', request=request)
        return Response({'count': len(records), 'ids': [r.id for r in records]}, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['put'])
    def update_remarks(self, request, pk=None):
        sf9 = self.get_object()
        sf9.adviser_remarks = request.data.get('adviser_remarks', sf9.adviser_remarks)
        sf9.principal_remarks = request.data.get('principal_remarks', sf9.principal_remarks)
        sf9.save(update_fields=['adviser_remarks', 'principal_remarks'])
        return Response(SchoolForm9DetailSerializer(sf9).data)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        sf9 = self.get_object()
        try:
            pdf_bytes = _generate_sf9_pdf(sf9)
        except ImportError:
            return Response({'error': 'reportlab required'}, status=500)
        log_audit_action(user=request.user, action='export_pdf', model_name='SchoolForm9',
                         object_id=sf9.id, description=f'Exported SF9 PDF: {sf9}', request=request)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="SF9_{sf9.school_year}_{sf9.student.last_name}_{sf9.grade_level}.pdf"'
        return response

    @action(detail=True, methods=['get'])
    def print_view(self, request, pk=None):
        return self.export_pdf(request, pk=pk)


def _generate_sf9_pdf(sf9):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LEGAL
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LEGAL,
        leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('C', parent=styles['Normal'], fontSize=7, leading=9)
    ss = ParagraphStyle('S', parent=styles['Normal'], fontSize=8, alignment=1)

    elements = []
    settings = SystemSetting.get_settings()
    elements.append(Paragraph("Republic of the Philippines", ss))
    elements.append(Paragraph("Department of Education", ss))
    elements.append(Paragraph(f"<b>{settings.site_name}</b>", ParagraphStyle('T', parent=styles['Title'], fontSize=11)))
    elements.append(Paragraph(f"School ID: {getattr(settings, 'school_id', '')}", ss))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"<b>SCHOOL FORM 9 (SF9) — Learner's Progress Report Card</b>",
        ParagraphStyle('FT', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=1, spaceAfter=4)))

    p = sf9.student.profile if hasattr(sf9.student, 'profile') else None
    info_data = [
        [Paragraph(f"<b>LRN:</b> {p.lrn if p else ''}", cs), '', Paragraph(f"<b>School Year:</b> {sf9.school_year}", cs), ''],
        [Paragraph(f"<b>Name:</b> {full_name(sf9.student)}", cs), '', Paragraph(f"<b>Grade:</b> {sf9.grade_level}", cs), ''],
        [Paragraph(f"<b>Sex:</b> {(p.sex or '').title() if p else ''}", cs), '', Paragraph(f"<b>Section:</b> {sf9.section}", cs), ''],
        [Paragraph(f"<b>Adviser:</b> {full_name(sf9.adviser) if sf9.adviser else ''}", cs), '', '', ''],
    ]
    info_table = Table(info_data, colWidths=[3*inch, 0.5*inch, 3*inch, 0.5*inch])
    info_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.append(info_table)
    elements.append(Spacer(1, 6))

    subjects = sf9.subjects.select_related('subject').order_by('subject__name')
    headers = ['Subject', 'Q1', 'Q2', 'Q3', 'Q4', 'Final', 'Remarks']
    table_data = [[Paragraph(f'<b>{h}</b>', cs) for h in headers]]
    for subj in subjects:
        table_data.append([
            Paragraph(subj.subject.name, cs),
            Paragraph(str(subj.q1) if subj.q1 else '—', cs),
            Paragraph(str(subj.q2) if subj.q2 else '—', cs),
            Paragraph(str(subj.q3) if subj.q3 else '—', cs),
            Paragraph(str(subj.q4) if subj.q4 else '—', cs),
            Paragraph(str(subj.final_rating) if subj.final_rating else '—', cs),
            Paragraph(subj.remarks, cs),
        ])

    table_data.append([
        Paragraph(f'<b>General Average: {sf9.general_average or "—"}</b>', cs), '', '', '', '',
        Paragraph(f'<b>Status: {sf9.promotion_status}</b>', cs), '',
    ])

    elements.append(Table(table_data, colWidths=[2.2*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.8*inch, 1.5*inch],
        repeatRows=1, style=TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D1B4D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])))

    elements.append(Spacer(1, 8))
    att_data = [
        [Paragraph(f"<b>Attendance:</b> Present: {sf9.days_present} | Absent: {sf9.days_absent} | Tardy: {sf9.days_tardy}", cs)],
        [Paragraph(f"<b>Adviser's Remarks:</b> {sf9.adviser_remarks or '—'}", cs)],
        [Paragraph(f"<b>Principal's Remarks:</b> {sf9.principal_remarks or '—'}", cs)],
    ]
    elements.append(Table(att_data, colWidths=[7*inch]))
    elements.append(Spacer(1, 12))
    sig_data = [
        [Paragraph('<b>Adviser:</b> ___________________', cs), '', Paragraph('<b>Principal:</b> ___________________', cs)],
        [Paragraph(f'{full_name(sf9.adviser) if sf9.adviser else ""}', cs), '', ''],
    ]
    elements.append(Table(sig_data, colWidths=[3*inch, 1*inch, 3*inch]))

    doc.build(elements)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# SF10 — Learner's Permanent Academic Record
# ═══════════════════════════════════════════════════════════════════════════════

class SchoolForm10ViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['student__last_name', 'student__first_name', 'student__profile__lrn']
    ordering_fields = ['generated_at', 'status']
    ordering = ['-generated_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return SchoolForm10ListSerializer
        return SchoolForm10DetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = SchoolForm10.objects.select_related('student', 'student__profile', 'generated_by').prefetch_related('academic_records__subjects__subject')
        if user.role == 'student':
            qs = qs.filter(student=user)
        elif user.role == 'parent':
            linked = user.profile.linked_students.all() if hasattr(user, 'profile') else User.objects.none()
            qs = qs.filter(student__in=linked)
        sf_status = self.request.query_params.get('status')
        if sf_status:
            qs = qs.filter(status=sf_status)
        return qs.distinct()

    def perform_destroy(self, instance):
        log_audit_action(user=self.request.user, action='delete', model_name='SchoolForm10',
                         object_id=instance.id, object_repr=str(instance),
                         description=f'Deleted SF10: {instance}', request=self.request)
        instance.delete()

    @action(detail=False, methods=['post'])
    def generate(self, request):
        serializer = GenerateSF10Serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        sf10 = serializer.save()
        log_audit_action(user=request.user, action='generate', model_name='SchoolForm10',
                         object_id=sf10.id, object_repr=str(sf10),
                         description=f'Generated SF10 for {full_name(sf10.student)}', request=request)
        return Response(SchoolForm10DetailSerializer(sf10).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def regenerate(self, request, pk=None):
        sf10 = self.get_object()
        data = {'student_id': sf10.student_id, 'regenerate': True}
        serializer = GenerateSF10Serializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        new_sf10 = serializer.save()
        log_audit_action(user=request.user, action='regenerate', model_name='SchoolForm10',
                         object_id=new_sf10.id, object_repr=str(new_sf10),
                         description=f'Regenerated SF10 for {full_name(new_sf10.student)}', request=request)
        return Response(SchoolForm10DetailSerializer(new_sf10).data)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        sf10 = self.get_object()
        try:
            pdf_bytes = _generate_sf10_pdf(sf10)
        except ImportError:
            return Response({'error': 'reportlab required'}, status=500)
        log_audit_action(user=request.user, action='export_pdf', model_name='SchoolForm10',
                         object_id=sf10.id, description=f'Exported SF10 PDF: {sf10}', request=request)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="SF10_{sf10.student.last_name}_{sf10.student.first_name}.pdf"'
        return response

    @action(detail=True, methods=['get'])
    def print_view(self, request, pk=None):
        return self.export_pdf(request, pk=pk)


def _generate_sf10_pdf(sf10):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LEGAL
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LEGAL,
        leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('C', parent=styles['Normal'], fontSize=7, leading=9)
    ss = ParagraphStyle('S', parent=styles['Normal'], fontSize=8, alignment=1)

    elements = []
    settings = SystemSetting.get_settings()
    p = sf10.student.profile if hasattr(sf10.student, 'profile') else None

    elements.append(Paragraph("Republic of the Philippines", ss))
    elements.append(Paragraph("Department of Education", ss))
    elements.append(Paragraph(f"<b>{settings.site_name}</b>", ParagraphStyle('T', parent=styles['Title'], fontSize=11)))
    elements.append(Paragraph(f"School ID: {getattr(settings, 'school_id', '')}", ss))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"<b>SCHOOL FORM 10 (SF10) — Learner's Permanent Academic Record</b>",
        ParagraphStyle('FT', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=1, spaceAfter=4)))

    info = [
        [Paragraph(f"<b>Name:</b> {full_name(sf10.student)}", cs), '', Paragraph(f"<b>LRN:</b> {p.lrn if p else ''}", cs), ''],
        [Paragraph(f"<b>Sex:</b> {(p.sex or '').title() if p else ''}", cs), '', Paragraph(f"<b>Birth Date:</b> {p.date_of_birth.strftime('%m/%d/%Y') if p and p.date_of_birth else ''}", cs), ''],
        [Paragraph(f"<b>School Years:</b> {sf10.school_year_from} — {sf10.school_year_to}", cs), '', '', ''],
    ]
    elements.append(Table(info, colWidths=[3*inch, 0.5*inch, 3*inch, 0.5*inch]))
    elements.append(Spacer(1, 6))

    for record in sf10.academic_records.select_related().order_by('order'):
        elements.append(Paragraph(
            f"<b>{record.school_year} — {record.grade_level} {record.section}</b>  |  "
            f"Gen. Avg: {record.general_average or '—'}  |  Status: {record.promotion_status}",
            ParagraphStyle('RY', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold',
                           backColor=colors.HexColor('#e9ecef'), spaceAfter=2, spaceBefore=6)))

        subjects = record.subjects.select_related('subject').order_by('subject__name')
        if subjects:
            headers = ['Subject', 'Q1', 'Q2', 'Q3', 'Q4', 'Final', 'Remarks']
            table_data = [[Paragraph(f'<b>{h}</b>', cs) for h in headers]]
            for subj in subjects:
                table_data.append([
                    Paragraph(subj.subject.name, cs),
                    Paragraph(str(subj.q1) if subj.q1 else '—', cs),
                    Paragraph(str(subj.q2) if subj.q2 else '—', cs),
                    Paragraph(str(subj.q3) if subj.q3 else '—', cs),
                    Paragraph(str(subj.q4) if subj.q4 else '—', cs),
                    Paragraph(str(subj.final_rating) if subj.final_rating else '—', cs),
                    Paragraph(subj.remarks, cs),
                ])
            elements.append(Table(table_data,
                colWidths=[2.2*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.8*inch, 1.5*inch],
                repeatRows=1, style=TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D1B4D')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ])))

        if record.awards:
            elements.append(Paragraph(f"<b>Awards:</b> {record.awards}", cs))
        if record.remarks:
            elements.append(Paragraph(f"<b>Remarks:</b> {record.remarks}", cs))

    doc.build(elements)
    return buf.getvalue()
