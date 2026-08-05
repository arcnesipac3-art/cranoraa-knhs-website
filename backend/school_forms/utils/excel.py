"""Excel export utilities for school forms - DepEd SF1 format.

Performance optimizations for large datasets:
- Write-only mode for large sheets (500+ rows)
- Batch cell writes to reduce object creation
- Optimized style application with style objects reuse
- Memory-efficient date formatting
"""
import io
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.writer.excel import save_virtual_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


def generate_excel_report(form_type, data):
    """Generate Excel report from form data."""
    if form_type == 'SF1':
        return _generate_sf1_xlsx(data)
    if XLSX_AVAILABLE:
        return _generate_generic_xlsx(form_type, data)
    return _generate_csv_fallback(form_type, data)


def _create_style_objects():
    """Create reusable style objects for performance optimization.
    
    Returns a dictionary of all styles used in SF1 generation.
    Reusing style objects instead of creating new ones for each cell
    significantly reduces memory usage and improves performance.
    """
    return {
        'header_font': Font(name='Arial', size=16, bold=True),
        'sub_header_font': Font(name='Arial', size=10, bold=True),
        'label_font': Font(name='Arial', size=9, bold=True),
        'value_font': Font(name='Arial', size=9),
        'small_font': Font(name='Arial', size=7),
        'small_bold_font': Font(name='Arial', size=7, bold=True),
        'title_font': Font(name='Arial', size=12, bold=True),
        'italic_gray_font': Font(name='Arial', size=7, italic=True, color='808080'),
        'italic_small_font': Font(name='Arial', size=8, italic=True),
        'center_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left_align': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        ),
        'header_fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'male_fill': PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid'),
        'female_fill': PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid'),
        'totals_fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    }


def _format_birthdate(birthdate):
    """Format birthdate efficiently for Excel export."""
    if birthdate:
        try:
            return birthdate.strftime('%m-%d-%Y')
        except (AttributeError, ValueError):
            pass
    return ''


def _write_student_row(ws, current_row, student, fill_style, styles, use_write_only=False):
    """Write a single student row efficiently.
    
    Args:
        ws: Worksheet object
        current_row: Row number to write to
        student: Student data dictionary
        fill_style: Fill pattern for the row
        styles: Style objects dictionary
        use_write_only: Whether using write-only mode
    """
    row_data = [
        student.get('no', ''),
        student.get('lrn', ''),
        student.get('name', ''),
        student.get('sex', ''),
        _format_birthdate(student.get('birthdate')),
        student.get('age', ''),
        student.get('mother_tongue', ''),
        student.get('indigenous_people', ''),
        student.get('religion', ''),
        student.get('house_number', ''),
        student.get('barangay', ''),
        student.get('city_municipality', ''),
        student.get('province', ''),
        student.get('father_name', ''),
        student.get('mother_name', ''),
        student.get('guardian_name', ''),
        student.get('guardian_relationship', ''),
        student.get('contact_number', ''),
        student.get('learning_modality', ''),
        student.get('remarks', ''),
    ]
    
    if use_write_only:
        # Write-only mode: create cells with styles
        from openpyxl.cell import WriteOnlyCell
        styled_row = []
        for value in row_data:
            cell = WriteOnlyCell(ws, value=value)
            cell.font = styles['small_font']
            cell.border = styles['thin_border']
            cell.alignment = styles['left_align']
            cell.fill = fill_style
            styled_row.append(cell)
        ws.append(styled_row)
    else:
        # Regular mode: write then style
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = styles['small_font']
            cell.border = styles['thin_border']
            cell.alignment = styles['left_align']
            cell.fill = fill_style


def _generate_sf1_xlsx(data):
    """Generate official DepEd SF1 School Register Excel.
    
    Optimized for large datasets (500+ students):
    - Write-only mode for sheets with 100+ students
    - Reusable style objects to reduce memory
    - Batch cell operations
    """
    if not XLSX_AVAILABLE:
        return _generate_csv_fallback('SF1', data)

    school_info = data.get('school_info', {})
    classrooms = data.get('classrooms', [])
    school_head = data.get('school_head_name', '')
    generated_date = data.get('generated_date', date.today().strftime('%B %d, %Y'))
    
    # Determine if we need write-only mode for performance
    total_students = sum(
        len(c.get('male_students', [])) + len(c.get('female_students', [])) 
        for c in classrooms
    )
    use_write_only = total_students > 100
    
    # Create workbook with appropriate mode
    wb = openpyxl.Workbook(write_only=use_write_only) if use_write_only else openpyxl.Workbook()

    # Create reusable style objects to reduce memory overhead
    styles = _create_style_objects()

    # Styles shorthand
    header_font = styles['header_font']
    sub_header_font = styles['sub_header_font']
    label_font = styles['label_font']
    value_font = styles['value_font']
    small_font = styles['small_font']
    small_bold_font = styles['small_bold_font']
    title_font = styles['title_font']
    center_align = styles['center_align']
    left_align = styles['left_align']
    thin_border = styles['thin_border']
    header_fill = styles['header_fill']
    male_fill = styles['male_fill']
    female_fill = styles['female_fill']
    totals_fill = styles['totals_fill']

    # SF1 Column headers matching the official template
    col_headers = [
        'No.', 'LRN', 'NAME\n(Last Name, First Name,\nMiddle Name)', 'SEX\n(M/F)',
        'BIRTH DATE\n(mm/dd/yyyy)', 'AGE\n(as of\n1st Fri\nof June)',
        'MOTHER\nTONGUE\n(Grade 1\nto 3 only)', 'IP/Ethnic\nGroup',
        'RELIGION', 'ADDRESS\n(House #/Street/\nSitio/Purok)', 'Barangay',
        'Municipality\n/ City', 'Province',
        "FATHER'S Name\n(Last Name, First Name,\nMiddle Name)",
        "MOTHER'S Maiden Name\n(Last Name, First Name,\nMiddle Name)",
        'GUARDIAN\n(if Not Parent)\nName', 'Relationship',
        'Contact\nNumber of\nParent/\nGuardian',
        'Learning\nModality', 'REMARKS'
    ]

    col_widths = [5, 14, 30, 6, 12, 8, 12, 10, 10, 18, 14, 14, 14, 28, 28, 22, 10, 12, 10, 12]

    for classroom_data in classrooms:
        classroom = classroom_data.get('classroom', {})
        male_students = classroom_data.get('male_students', [])
        female_students = classroom_data.get('female_students', [])
        total_male = classroom_data.get('total_male', 0)
        total_female = classroom_data.get('total_female', 0)
        total_combined = classroom_data.get('total_combined', 0)
        remarks_counts = classroom_data.get('remarks_counts', {})

        section_name = classroom.get('section', 'Section')
        grade_level = classroom.get('grade_level', '')
        adviser_name = classroom.get('adviser', '')
        academic_year = classroom.get('academic_year', '')

        ws = wb.create_sheet(title=section_name[:31])

        # ── HEADER SECTION ──────────────────────────────────────────────
        ws.merge_cells('A1:T1')
        ws['A1'] = 'School Form 1 (SF 1) School Register'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:T2')
        ws['A2'] = '(This replaces Form 1 Master List & SF1-Form 2-Family Background and Profile)'
        ws['A2'].font = Font(name='Arial', size=8, italic=True)
        ws['A2'].alignment = center_align

        # Row 3: School ID, Region, Division
        row = 3
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'School ID'
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].border = thin_border
        ws[f'C{row}'] = school_info.get('school_id', '')
        ws[f'C{row}'].font = value_font
        ws[f'C{row}'].border = thin_border

        ws[f'D{row}'] = 'Region'
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].border = thin_border
        ws.merge_cells(f'E{row}:G{row}')
        ws[f'E{row}'] = school_info.get('region', '')
        ws[f'E{row}'].font = value_font
        ws[f'E{row}'].border = thin_border

        ws[f'H{row}'] = 'Division'
        ws[f'H{row}'].font = label_font
        ws[f'H{row}'].border = thin_border
        ws.merge_cells(f'I{row}:K{row}')
        ws[f'I{row}'] = school_info.get('division', '')
        ws[f'I{row}'].font = value_font
        ws[f'I{row}'].border = thin_border

        # Row 4: School Name
        row = 4
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'School Name'
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].border = thin_border
        ws.merge_cells(f'C{row}:K{row}')
        ws[f'C{row}'] = school_info.get('school_name', '')
        ws[f'C{row}'].font = value_font
        ws[f'C{row}'].border = thin_border

        # Row 5: School Year, Grade Level, Section
        row = 5
        ws[f'A{row}'] = 'School Year'
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].border = thin_border
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = academic_year
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].border = thin_border

        ws[f'E{row}'] = 'Grade Level'
        ws[f'E{row}'].font = label_font
        ws[f'E{row}'].border = thin_border
        ws.merge_cells(f'F{row}:G{row}')
        ws[f'F{row}'] = grade_level
        ws[f'F{row}'].font = value_font
        ws[f'F{row}'].border = thin_border

        ws[f'H{row}'] = 'Section'
        ws[f'H{row}'].font = label_font
        ws[f'H{row}'].border = thin_border
        ws.merge_cells(f'I{row}:K{row}')
        ws[f'I{row}'] = section_name
        ws[f'I{row}'].font = value_font
        ws[f'I{row}'].border = thin_border

        # Row 6: Column headers
        row = 6
        if not use_write_only:
            for col_idx, header in enumerate(col_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = small_bold_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            # Set column widths
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Row heights
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 15
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[5].height = 20
            ws.row_dimensions[6].height = 55

        # ── MALE STUDENTS ───────────────────────────────────────────────
        current_row = 7
        if male_students:
            for student in male_students:
                _write_student_row(ws, current_row, student, male_fill, styles, use_write_only)
                if not use_write_only:
                    current_row += 1

        # TOTAL MALE row
        if not use_write_only:
            ws.merge_cells(f'A{current_row}:T{current_row}')
            ws[f'A{current_row}'] = f'{total_male} .... TOTAL MALE'
            ws[f'A{current_row}'].font = small_bold_font
            ws[f'A{current_row}'].alignment = left_align
            for col in range(1, 21):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).fill = totals_fill
        else:
            # Write-only mode: create styled row
            from openpyxl.cell import WriteOnlyCell
            total_row = []
            for col_idx in range(1, 21):
                cell = WriteOnlyCell(ws, value=f'{total_male} .... TOTAL MALE' if col_idx == 1 else '')
                cell.font = small_bold_font
                cell.border = thin_border
                cell.alignment = left_align
                cell.fill = totals_fill
                total_row.append(cell)
            ws.append(total_row)
        current_row += 1

        # ── FEMALE STUDENTS ─────────────────────────────────────────────
        if female_students:
            for student in female_students:
                _write_student_row(ws, current_row, student, female_fill, styles, use_write_only)
                if not use_write_only:
                    current_row += 1

        # TOTAL FEMALE row
        if not use_write_only:
            ws.merge_cells(f'A{current_row}:T{current_row}')
            ws[f'A{current_row}'] = f'{total_female} .... TOTAL FEMALE'
            ws[f'A{current_row}'].font = small_bold_font
            ws[f'A{current_row}'].alignment = left_align
            for col in range(1, 21):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).fill = totals_fill
        else:
            from openpyxl.cell import WriteOnlyCell
            total_row = []
            for col_idx in range(1, 21):
                cell = WriteOnlyCell(ws, value=f'{total_female} .... TOTAL FEMALE' if col_idx == 1 else '')
                cell.font = small_bold_font
                cell.border = thin_border
                cell.alignment = left_align
                cell.fill = totals_fill
                total_row.append(cell)
            ws.append(total_row)
        current_row += 1

        # COMBINED TOTAL row
        if not use_write_only:
            ws.merge_cells(f'A{current_row}:T{current_row}')
            ws[f'A{current_row}'] = f'{total_combined} .... COMBINED'
            ws[f'A{current_row}'].font = small_bold_font
            ws[f'A{current_row}'].alignment = left_align
            for col in range(1, 21):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).fill = totals_fill
        else:
            from openpyxl.cell import WriteOnlyCell
            total_row = []
            for col_idx in range(1, 21):
                cell = WriteOnlyCell(ws, value=f'{total_combined} .... COMBINED' if col_idx == 1 else '')
                cell.font = small_bold_font
                cell.border = thin_border
                cell.alignment = left_align
                cell.fill = totals_fill
                total_row.append(cell)
            ws.append(total_row)
        current_row += 1

        # ── REMARKS LEGEND ──────────────────────────────────────────────
        current_row += 1
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = 'List and Code of Indicators under REMARKS column'
        ws[f'A{current_row}'].font = small_bold_font
        current_row += 1

        remarks_legend = [
            ['Indicator', 'Code', 'Required Information', '', 'Indicator', 'Code', 'Required Information'],
            ['Transferred Out', 'TrnO', "Name of Public (P) Private (PR) School\n& Effectivity Date", '', 'CCT Recipient', 'CCT', 'CCT Control/Reference number &\nEffectivity Date'],
            ['Transferred In', 'TrnI', 'Reason and Effectivity Date', '', 'Balik-Aral', 'BA', "Name of school last attended & Year\nof Drop Out (if applicable)"],
            ['Dropped', 'DR', 'Reason (Enrollment beyond 1st Friday)', '', 'Special Needs Education\nAccelerated', 'SNED', 'Specify Level & Effectivity Data'],
        ]
        for legend_row in remarks_legend:
            for col_idx, val in enumerate(legend_row, 1):
                ws.cell(row=current_row, column=col_idx, value=val).font = small_font
                ws.cell(row=current_row, column=col_idx).border = thin_border
                ws.cell(row=current_row, column=col_idx).alignment = left_align
            current_row += 1

        # ── SIGNATURE SECTION ───────────────────────────────────────────
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = 'Prepared by:'
        ws[f'A{current_row}'].font = small_bold_font
        current_row += 2
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = adviser_name
        ws[f'A{current_row}'].font = small_bold_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = '(Signature of Adviser over Printed Name)'
        ws[f'A{current_row}'].font = small_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')

        current_row += 2
        ws.merge_cells(f'H{current_row}:Q{current_row}')
        ws[f'H{current_row}'] = 'Certified Correct:'
        ws[f'H{current_row}'].font = small_bold_font
        current_row += 2
        ws.merge_cells(f'H{current_row}:Q{current_row}')
        ws[f'H{current_row}'] = school_head
        ws[f'H{current_row}'].font = small_bold_font
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        ws.merge_cells(f'H{current_row}:Q{current_row}')
        ws[f'H{current_row}'] = '(Signature of School Head over Printed Name)'
        ws[f'H{current_row}'].font = small_font
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')

        # ── FOOTER ──────────────────────────────────────────────────────
        current_row += 2
        if not use_write_only:
            ws.merge_cells(f'A{current_row}:T{current_row}')
            ws[f'A{current_row}'] = f'Generated on: {generated_date}'
            ws[f'A{current_row}'].font = styles['italic_gray_font']
            ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
            
            current_row += 1
            ws.merge_cells(f'A{current_row}:T{current_row}')
            ws[f'A{current_row}'] = 'This is an official DepEd School Form 1 (SF1) - School Register'
            ws[f'A{current_row}'].font = styles['italic_gray_font']
            ws[f'A{current_row}'].alignment = center_align

            # Set print area and page setup
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.horizontalCentered = True
            ws.page_setup.verticalCentered = False
            
            # Set print margins (in inches)
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3
            
            # Freeze panes at header row
            ws.freeze_panes = 'A7'
        else:
            # Write-only mode: just append footer rows
            from openpyxl.cell import WriteOnlyCell
            footer_row1 = []
            for col_idx in range(1, 21):
                cell = WriteOnlyCell(ws, value=f'Generated on: {generated_date}' if col_idx == 1 else '')
                cell.font = styles['italic_gray_font']
                footer_row1.append(cell)
            ws.append(footer_row1)
            
            footer_row2 = []
            for col_idx in range(1, 21):
                cell = WriteOnlyCell(ws, value='This is an official DepEd School Form 1 (SF1) - School Register' if col_idx == 1 else '')
                cell.font = styles['italic_gray_font']
                footer_row2.append(cell)
            ws.append(footer_row2)

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _generate_generic_xlsx(form_type, data):
    """Generic XLSX generation for other forms."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = form_type
    ws.append([form_type])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _generate_csv_fallback(form_type, data):
    """CSV fallback when openpyxl is not available."""
    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    if form_type == 'SF1':
        classrooms = data.get('classrooms', []) if isinstance(data, dict) else data
        writer.writerow(['No', 'LRN', 'Name', 'Sex', 'Birthdate', 'Age', 'Mother Tongue', 'IP', 'Religion',
                         'House #', 'Barangay', 'City/Municipality', 'Province',
                         "Father's Name", "Mother's Name", 'Guardian', 'Relationship', 'Contact', 'Modality', 'Remarks'])
        for classroom_data in classrooms:
            for student in classroom_data.get('male_students', []) + classroom_data.get('female_students', []):
                bd = student.get('birthdate')
                writer.writerow([
                    student.get('no', ''), student.get('lrn', ''), student.get('name', ''),
                    student.get('sex', ''), bd.strftime('%m-%d-%Y') if bd else '', student.get('age', ''),
                    student.get('mother_tongue', ''), student.get('indigenous_people', ''),
                    student.get('religion', ''), student.get('house_number', ''),
                    student.get('barangay', ''), student.get('city_municipality', ''),
                    student.get('province', ''), student.get('father_name', ''),
                    student.get('mother_name', ''), student.get('guardian_name', ''),
                    student.get('guardian_relationship', ''), student.get('contact_number', ''),
                    student.get('learning_modality', ''), student.get('remarks', ''),
                ])

    return output.getvalue().encode('utf-8-sig')
